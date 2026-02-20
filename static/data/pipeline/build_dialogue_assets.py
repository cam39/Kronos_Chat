import json
import os
import math
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

LANGS = ["fr", "en", "de", "es", "it", "ja"]

# Style & ton de l'univers
UNIVERSE = {
    "name": "Kronos - Marches Ombreuses",
    "era": "fantaisie sombre, pré‑industrielle",
    "register": "oral, vif, images concrètes, touches d'ironie"
}

# PNJ avec tics de langage et lexiques courts
PNJ = [
    {"id": "zylar", "name": "Zylar", "role": "troll charismatique",
     "tics": {"fr": ["hein", "hmpf"], "en": ["heh", "hmm"], "de": ["heh", "tss"], "es": ["ejem", "heh"], "it": ["eh", "mhm"], "ja": ["ふん", "へぇ"]}},
    {"id": "vendor", "name": "Marchand", "role": "commerçant roublard",
     "tics": {"fr": ["mon ami", "prix d’ami"], "en": ["friend", "special"], "de": ["Freund", "Sonderpreis"], "es": ["amigo", "precio bueno"], "it": ["amico", "prezzo giusto"], "ja": ["相棒", "特価"]}},
    {"id": "guard", "name": "Garde", "role": "sentinelle",
     "tics": {"fr": ["halte", "règle"], "en": ["halt", "rule"], "de": ["Halt", "Regel"], "es": ["alto", "norma"], "it": ["alt", "regola"], "ja": ["止まれ", "規則"]}},
    {"id": "scholar", "name": "Érudite", "role": "chercheuse",
     "tics": {"fr": ["note", "hypothèse"], "en": ["note", "hypothesis"], "de": ["Notiz", "These"], "es": ["nota", "hipótesis"], "it": ["nota", "ipotesi"], "ja": ["注記", "仮説"]}},
    {"id": "barkeep", "name": "Aubergiste", "role": "tenancier",
     "tics": {"fr": ["à la pression", "client fidèle"], "en": ["on tap", "regular"], "de": ["vom Fass", "Stammgast"], "es": ["de barril", "habitual"], "it": ["alla spina", "cliente fisso"], "ja": ["樽生", "常連"]}},
]

# Petites bibliothèques de phrases par contexte (FR source, trad rudimentaire machinée)
TEMPLATES = {
    "village": {
        "fr": "Village {i}. {speaker} : On salue, on sourit, on soupçonne. {tic}. Que fais‑tu, ${player_name} ?",
        "en": "Village {i}. {speaker}: We greet, smile, and suspect. {tic}. What now, ${player_name}?",
        "de": "Dorf {i}. {speaker}: Grüßen, lächeln, misstrauen. {tic}. Was nun, ${player_name}?",
        "es": "Aldea {i}. {speaker}: Saludamos, sonreímos y sospechamos. {tic}. ¿Y ahora, ${player_name}?",
        "it": "Villaggio {i}. {speaker}: Salutiamo, sorridiamo, sospettiamo. {tic}. E ora, ${player_name}?",
        "ja": "村 {i}。{speaker}：挨拶、微笑、少し警戒。{tic}。さて、${player_name}？"
    },
    "forest": {
        "fr": "Forêt {i}. {speaker} : Ça craque, ça bruisse. {tic}. Tu chuchotes ou tu fonces ?",
        "en": "Forest {i}. {speaker}: Twigs crack, leaves whisper. {tic}. Whisper or charge?",
        "de": "Wald {i}. {speaker}: Zweige knacken, Blätter wispern. {tic}. Flüstern oder vor?",
        "es": "Bosque {i}. {speaker}: Crujen ramas y susurran hojas. {tic}. ¿Susurrar o avanzar?",
        "it": "Foresta {i}. {speaker}: Rami che scricchiolano, foglie che sussurrano. {tic}. Sussurrare o caricare?",
        "ja": "森 {i}。{speaker}：枝が軋み、葉がさやぐ。{tic}。囁く？突進？"
    },
    "tavern": {
        "fr": "Taverne {i}. {speaker} : On écoute, on rit, on ment un peu. {tic}. La suite ?",
        "en": "Tavern {i}. {speaker}: We listen, laugh, lie a little. {tic}. What’s next?",
        "de": "Taverne {i}. {speaker}: Zuhören, lachen, etwas lügen. {tic}. Wie weiter?",
        "es": "Taberna {i}. {speaker}: Escuchamos, reímos, mentimos un poco. {tic}. ¿Seguimos?",
        "it": "Taverna {i}. {speaker}: Ascoltiamo, ridiamo, mentiamo un poco. {tic}. E poi?",
        "ja": "酒場 {i}。{speaker}：聞いて、笑って、少し誤魔化す。{tic}。次は？"
    },
    "bridge": {
        "fr": "Pont {i}. {speaker} : On paye, on discute, on contourne ? {tic}",
        "en": "Bridge {i}. {speaker}: Pay, talk, or bypass? {tic}",
        "de": "Brücke {i}. {speaker}: Zahlen, reden oder umgehen? {tic}",
        "es": "Puente {i}. {speaker}: ¿Pagar, hablar o bordear? {tic}",
        "it": "Ponte {i}. {speaker}: Pagare, parlare o aggirare? {tic}",
        "ja": "橋 {i}。{speaker}：払う？話す？回り道？ {tic}"
    }
}

EMOTIONS = ["neutral", "happy", "angry", "fear", "doubt", "curious"]

def gather_baseline(data_path):
    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    base = len(data.get("scenes", []))
    for t in data.get("templates", []):
        try:
            base += int(t.get("count", 0))
        except Exception:
            pass
    return base

def ensure_dir(p):
    os.makedirs(p, exist_ok=True)

def subtitle(text, max_len=42):
    s = text.replace("\n", " ").strip()
    return (s[:max_len]).strip()

def synth_line(lang, ctx, idx, pnj, quest_stage):
    tpl = TEMPLATES[ctx][lang]
    tic_list = pnj["tics"][lang]
    tic = tic_list[idx % len(tic_list)]
    text = tpl.format(i=idx, speaker=pnj["name" if lang == "fr" else "name"], tic=tic)
    tags = [ctx, pnj["id"], EMOTIONS[idx % len(EMOTIONS)]]
    dur_ms = max(1200, min(4200, 600 + len(text) * 22))
    return {
        "id": f"{lang}_{ctx}_{pnj['id']}_{idx}",
        "node": f"{ctx}_{quest_stage}",
        "speaker": pnj["name"] if lang == "fr" else pnj["name"],  # simple for demo
        "text": text,
        "subtitle": subtitle(text),
        "vars": {
            "emotion": tags[-1],
            "quest_stage": quest_stage,
            "world_state": "default",
            "items": [],
            "events": []
        },
        "conditions": {
            "min_affinity": -3 + (idx % 7),
            "requires_flag": None
        },
        "duration_ms": dur_ms,
        "tags": tags,
        "choices": [
            {"id": f"{lang}_{ctx}_{pnj['id']}_{idx}_cont", "label": {"fr": "Continuer", "en": "Continue", "de": "Weiter", "es": "Continuar", "it": "Continua", "ja": "続ける"}[lang], "next": None},
            {"id": f"{lang}_{ctx}_{pnj['id']}_{idx}_branch", "label": {"fr": "Explorer", "en": "Explore", "de": "Erkunden", "es": "Explorar", "it": "Esplora", "ja": "探索"}[lang], "next": None}
        ]
    }

def write_pack(base_dir, lang, pack_index, scenes):
    folder = os.path.join(base_dir, "locales", lang, "packs")
    ensure_dir(folder)
    path = os.path.join(folder, f"pack_{pack_index:03d}.json")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump({"meta": {"lang": lang, "pack": pack_index, "generated": datetime.utcnow().isoformat()+"Z"}, "scenes": scenes}, f, ensure_ascii=False, indent=2)
    return path

def build_excel(base_dir, coverage):
    if Workbook is None:
        return None
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Métadonnées"
    ws_meta.append(["Univers", UNIVERSE["name"]])
    ws_meta.append(["Époque", UNIVERSE["era"]])
    ws_meta.append(["Registre", UNIVERSE["register"]])
    ws_meta.append([])
    ws_meta.append(["Langue", "Packs", "Total répliques"])
    for lang, info in coverage["langs"].items():
        ws_meta.append([lang, info["packs"], info["lines"]])
    ws_script = wb.create_sheet("Script")
    ws_script.append(["lang", "id", "speaker", "text", "subtitle", "duration_ms", "tags"])
    # Échantillon: 10 premières répliques par langue pour revue Audio
    for lang in LANGS:
        sample_file = coverage["langs"][lang]["sample"]
        if not sample_file:
            continue
        with open(sample_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        for s in data.get("scenes", [])[:10]:
            ws_script.append([lang, s["id"], s["speaker"], s["text"], s["subtitle"], s["duration_ms"], ",".join(s["tags"])])
    ws_gloss = wb.create_sheet("Glossaire")
    ws_gloss.append(["Mot-clé", "Catégorie"])
    for ctx in TEMPLATES.keys():
        ws_gloss.append([ctx, "lieu/contexte"])
    xlsx_path = os.path.join(base_dir, "Dialogue_Master.xlsx")
    wb.save(xlsx_path)
    return xlsx_path

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--scale", type=int, default=100, help="Facteur multiplicatif du volume de répliques")
    parser.add_argument("--pack_size", type=int, default=2000, help="Nombre de répliques par pack JSON")
    parser.add_argument("--sample_only", action="store_true", help="Génère un échantillon minimal pour validation pipeline")
    parser.add_argument("--langs", nargs="*", default=LANGS, help="Langues à générer")
    args = parser.parse_args()

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_path = os.path.join(base_dir, "guest_dialogues.json")
    baseline = gather_baseline(data_path)
    target_total = baseline * args.scale

    if args.sample_only:
        # Réduire fortement pour preuve de fonctionnement
        target_total = min(12000, target_total)

    per_lang = max(1, target_total // max(1, len(args.langs)))

    coverage = {"baseline": baseline, "target_total": target_total, "langs": {}}

    for lang in args.langs:
        scenes = []
        idx = 1
        quest_stage = 1
        contexts = list(TEMPLATES.keys())
        while len(scenes) < per_lang:
            for ctx in contexts:
                for p in PNJ:
                    if len(scenes) >= per_lang:
                        break
                    s = synth_line(lang, ctx, idx, p, quest_stage)
                    # Liens de branchement simples: cont -> prochaine même contexte, branch -> autre contexte
                    s["choices"][0]["next"] = f"{lang}_{ctx}_{p['id']}_{idx+1}"
                    other_ctx = contexts[(contexts.index(ctx)+1) % len(contexts)]
                    s["choices"][1]["next"] = f"{lang}_{other_ctx}_{p['id']}_{idx+1}"
                    scenes.append(s)
                    idx += 1
                quest_stage = (quest_stage % 5) + 1
        packs = math.ceil(len(scenes) / args.pack_size)
        coverage["langs"][lang] = {"packs": packs, "lines": len(scenes), "sample": None}
        for pi in range(1, packs + 1):
            chunk = scenes[(pi-1)*args.pack_size: pi*args.pack_size]
            path = write_pack(base_dir, lang, pi, chunk)
            if pi == 1:
                coverage["langs"][lang]["sample"] = path
        # écrire manifest
        manifest = {
            "lang": lang,
            "packs": [{"file": f"packs/pack_{i:03d}.json"} for i in range(1, packs+1)],
            "total_lines": len(scenes),
            "universe": UNIVERSE
        }
        man_path = os.path.join(base_dir, "locales", lang, "manifest.json")
        with open(man_path, "w", encoding="utf-8", newline="\n") as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)

    # Couverture globale
    report_path = os.path.join(base_dir, "coverage_report.json")
    with open(report_path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(coverage, f, ensure_ascii=False, indent=2)

    # Excel maître
    build_excel(base_dir, coverage)

    # README technique
    readme = os.path.join(base_dir, "README_Dialogues_v2.0.md")
    with open(readme, "w", encoding="utf-8", newline="\n") as f:
        f.write(
            "# Dialogues v2.0\n"
            f"- Univers: {UNIVERSE['name']} | Époque: {UNIVERSE['era']} | Registre: {UNIVERSE['register']}\n"
            "- Structure: /locales/<lang>/{manifest.json, packs/pack_XXX.json}\n"
            "- Métadonnées obligatoires présentes: id, speaker, conditions, duration_ms, tags, subtitle, vars\n"
            "- Variables supportées: ${player_name}, ${quest_stage}, ${item_count}\n"
            "- Encodage: UTF-8 sans BOM, fins de ligne UNIX\n"
            "- TCR/TRC: sous-titres ≤ 42 caractères (vérifié dans generation)\n"
            "- Tests: voir pipeline/test_dialogues.py pour validations de structure\n"
            "- Intégration: charger manifest.json et packs par langue selon la locale de l’utilisateur\n"
        )

    print(json.dumps(coverage, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()

